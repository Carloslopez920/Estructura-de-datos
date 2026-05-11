# file_utils.py
import os
import json
import csv
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
import re
from typing import List, Any, Tuple, Optional

# Verificar dependencias opcionales
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

class FileProcessor:
    """Clase para procesar diferentes tipos de archivos"""
    
    @staticmethod
    def get_supported_formats() -> List[Tuple[str, str]]:
        """Obtener lista de formatos soportados"""
        return [
            ("Todos los archivos soportados", "*.txt *.csv *.json *.xml *.yaml *.yml *.log *.dat *.xlsx *.xls"),
            ("📝 Texto", "*.txt *.log *.dat"),
            ("📊 Datos", "*.csv *.json *.xml *.yaml *.yml"),
            ("📈 Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    
    @staticmethod
    def read_file_raw(filepath: str) -> str:
        """Leer archivo y devolver contenido crudo según extensión"""
        extension = Path(filepath).suffix.lower()
        contenido = ""
        
        try:
            if extension in ['.txt', '.log', '.dat', '.ini', '.cfg', 
                           '.py', '.java', '.cpp', '.c', '.js', '.html', '.css', '.sql']:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    contenido = f.read()
            
            elif extension == '.csv':
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        contenido += " ".join(row) + "\n"
            
            elif extension == '.json':
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    contenido = json.dumps(data, ensure_ascii=False)
            
            elif extension == '.xml':
                tree = ET.parse(filepath)
                root = tree.getroot()
                contenido = ET.tostring(root, encoding='unicode')
            
            elif extension in ['.yaml', '.yml'] and YAML_AVAILABLE:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    contenido = str(data)
            
            elif extension in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
                # Para Excel se maneja de forma especial
                wb = load_workbook(filepath, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    contenido += f"\n--- HOJA: {sheet_name} ---\n"
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                contenido += str(cell.value) + " "
                    contenido += "\n"
                wb.close()
            
            else:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    contenido = f.read()
            
            return contenido
            
        except Exception as e:
            raise Exception(f"Error leyendo archivo: {str(e)}")
    
    @staticmethod
    def get_excel_sheets(filepath: str) -> List[str]:
        """Obtener nombres de hojas de un archivo Excel"""
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl no está instalado")
        
        wb = load_workbook(filepath, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    
    @staticmethod
    def process_excel_file(filepath: str, selected_sheets: List[str]) -> Tuple[str, List[Any]]:
        """Procesar archivo Excel con las hojas seleccionadas"""
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl no está instalado")
        
        wb = load_workbook(filepath, data_only=True)
        all_data = []
        content_raw = ""
        
        for sheet_name in selected_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content_raw += f"\n--- HOJA: {sheet_name} ---\n"
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            value = str(cell.value)
                            content_raw += value + " "
                            all_data.append(value)
                content_raw += "\n"
        
        wb.close()
        return content_raw, all_data
    
    @staticmethod
    def save_result(filepath: str, data: List[Any], metadata: dict) -> bool:
        """Guardar resultado en diferentes formatos"""
        try:
            extension = Path(filepath).suffix.lower()
            
            if extension == '.txt':
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write("="*70 + "\n")
                    f.write(f"RESULTADO DE COMBINACIÓN Y ORDENAMIENTO\n")
                    f.write(f"Fecha: {metadata.get('fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}\n")
                    f.write(f"Tipo de dato: {metadata.get('tipo_dato', 'unknown')}\n")
                    f.write("="*70 + "\n\n")
                    
                    for key, value in metadata.items():
                        if key not in ['fecha', 'tipo_dato']:
                            f.write(f"{key}: {value}\n")
                    
                    f.write("\nDATOS ORDENADOS:\n")
                    f.write("-"*40 + "\n")
                    for i, elem in enumerate(data, 1):
                        f.write(f"{i}. {elem}\n")
            
            elif extension == '.csv':
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Indice", "Valor"])
                    for i, elem in enumerate(data, 1):
                        writer.writerow([i, elem])
            
            elif extension == '.json':
                with open(filepath, 'w', encoding='utf-8') as f:
                    serializable_data = [str(elem) if not isinstance(elem, (int, float)) else elem for elem in data]
                    json.dump({
                        "metadata": metadata,
                        "datos_ordenados": serializable_data
                    }, f, indent=2, ensure_ascii=False)
            
            elif extension == '.html':
                FileProcessor._save_as_html(filepath, data, metadata)
            
            else:
                # Por defecto guardar como texto
                with open(filepath + '.txt', 'w', encoding='utf-8') as f:
                    for i, elem in enumerate(data, 1):
                        f.write(f"{i}. {elem}\n")
            
            return True
            
        except Exception as e:
            raise Exception(f"Error guardando archivo: {str(e)}")
    
    @staticmethod
    def _save_as_html(filepath: str, data: List[Any], metadata: dict):
        """Guardar resultado como HTML"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Resultados Combinados</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
.container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; }}
h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
h2 {{ color: #34495e; margin-top: 20px; }}
.stats {{ background: #ecf0f1; padding: 15px; border-radius: 5px; margin: 10px 0; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
th {{ background-color: #2c3e50; color: white; }}
tr:nth-child(even) {{ background-color: #f2f2f2; }}
.scrollable {{ max-height: 400px; overflow-y: auto; }}
</style>
</head>
<body>
<div class="container">
<h1>📊 Resultados de Combinación y Ordenamiento</h1>
<p><strong>Fecha:</strong> {metadata.get('fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}</p>
<p><strong>Tipo de dato:</strong> {metadata.get('tipo_dato', 'unknown')}</p>

<h2>📁 Información</h2>
<div class="stats">
""")
            
            for key, value in metadata.items():
                if key not in ['fecha', 'tipo_dato']:
                    f.write(f"<p><strong>{key}:</strong> {value}</p>\n")
            
            f.write("""</div>

<h2>🔢 Datos Ordenados</h2>
<div class="scrollable">
<table>
<tr><th>Posición</th><th>Valor</th></tr>
""")
            
            for i, elem in enumerate(data[:500], 1):
                f.write(f"<tr><td>{i}</td><td>{elem}</td></tr>\n")
            
            if len(data) > 500:
                f.write(f"<tr><td colspan='2'>... y {len(data)-500} elementos más ...</td></tr>\n")
            
            f.write("""</table>
</div>
</div>
</body>
</html>""")