"""
DEMO COMPLETA DE ORDENAMIENTOS
Ejecuta y visualiza todos los métodos de ordenamiento (Internos y Externos)
Soporta carga de archivos: TXT, XLSX y JSON
"""

import tkinter as tk
from tkinter import ttk, messagebox
import random
import time
import threading
import re
import json
from datetime import datetime

# Importar las librerías creadas
from ordenamiento_interno import OrdenamientoInterno
from ordenamiento_externo import OrdenamientoExterno

# Intentar importar openpyxl para soporte Excel
try:
    from openpyxl import load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


class DemoCompletaOrdenamiento:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title("Demo Completa - Métodos de Ordenamiento")
        self.ventana.geometry("1400x850")
        self.ventana.configure(bg='#1a1a2e')
        
        # Instancias de las librerías
        self.ordenador_interno = OrdenamientoInterno()
        self.ordenador_externo = OrdenamientoExterno()
        
        # Datos
        self.datos_actuales = []
        self.animacion_activa = False
        self.nombre_archivo_cargado = ""
        self.modo_visualizacion = "grafico"
        
        # Almacenar últimos resultados
        self.ultimos_resultados = {}
        
        # Métodos disponibles (ahora usando las instancias)
        self.metodos_internos = {
            "Burbuja": self.ordenador_interno.burbuja,
            "Inserción": self.ordenador_interno.insercion,
            "Selección": self.ordenador_interno.seleccion,
            "Shellsort": self.ordenador_interno.shell_sort,
            "Quicksort": self.ordenador_interno.quick_sort,
            "Heapsort": self.ordenador_interno.heap_sort,
            "Radixsort": self.ordenador_interno.radix_sort
        }
        
        self.metodos_externos = {
            "Intercalación": self._intercalacion_wrapper,
            "Mezcla Directa": self._mezcla_directa_wrapper,
            "Mezcla Equilibrada": self._mezcla_equilibrada_wrapper
        }
        
        self.setup_ui()
    
    def _intercalacion_wrapper(self, datos, callback=None):
        """Wrapper para intercalación (necesita dos listas)"""
        # Dividir datos en dos partes para demostrar intercalación
        mitad = len(datos) // 2
        lista1 = sorted(datos[:mitad])
        lista2 = sorted(datos[mitad:])
        resultado = self.ordenador_externo.intercalacion(lista1, lista2)
        return 0, resultado
    
    def _mezcla_directa_wrapper(self, datos, callback=None):
        """Wrapper para mezcla directa"""
        tamano_bloque = max(2, len(datos) // 5)
        resultado = self.ordenador_externo.mezcla_directa(datos, tamano_bloque)
        return 0, resultado
    
    def _mezcla_equilibrada_wrapper(self, datos, callback=None):
        """Wrapper para mezcla equilibrada"""
        tamano_bloque = max(2, len(datos) // 5)
        resultado = self.ordenador_externo.mezcla_equilibrada(datos, tamano_bloque)
        return 0, resultado
    
    def setup_ui(self):
        # Estilo
        style = ttk.Style()
        style.configure("TLabel", background="#1a1a2e", foreground="white")
        style.configure("TFrame", background="#1a1a2e")
        style.configure("TLabelframe", background="#1a1a2e", foreground="white")
        
        # Frame principal con paneles divididos
        paned_window = tk.PanedWindow(self.ventana, orient=tk.VERTICAL, bg="#1a1a2e", 
                                       sashrelief=tk.RAISED, sashwidth=5)
        paned_window.pack(fill=tk.BOTH, expand=True)
        
        # Panel superior para controles
        top_panel = tk.Frame(paned_window, bg="#1a1a2e", height=200)
        paned_window.add(top_panel, height=200)
        
        # Panel inferior para resultados (scrollable)
        bottom_panel = tk.Frame(paned_window, bg="#1a1a2e")
        paned_window.add(bottom_panel)
        
        # ========== PANEL SUPERIOR ==========
        # Título
        titulo = tk.Label(top_panel, text="📊 MÉTODOS DE ORDENAMIENTO - COMPARATIVA VISUAL",
                         font=("Arial", 16, "bold"), bg="#1a1a2e", fg="#00d4ff")
        titulo.pack(pady=5)
        
        # Frame de datos (primera fila)
        datos_frame = ttk.LabelFrame(top_panel, text="Configuración de Datos", padding="10")
        datos_frame.pack(fill=tk.X, pady=5, padx=10)
        
        # Botones de datos - Fila 1
        btn_frame1 = tk.Frame(datos_frame, bg="#1a1a2e")
        btn_frame1.pack(pady=5)
        
        tk.Button(btn_frame1, text="🎲 Datos Aleatorios", command=self.generar_aleatorios,
                 bg="#27ae60", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="✏️ Ingresar Manual", command=self.ingresar_manual,
                 bg="#2980b9", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📁 Cargar TXT", command=self.cargar_archivo_txt,
                 bg="#8e44ad", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📊 Cargar Excel", command=self.cargar_archivo_excel,
                 bg="#e67e22", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📄 Cargar JSON", command=self.cargar_archivo_json,
                 bg="#f39c12", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        
        # Botones de control - Fila 2
        btn_frame2 = tk.Frame(datos_frame, bg="#1a1a2e")
        btn_frame2.pack(pady=5)
        
        tk.Button(btn_frame2, text="▶️ Ejecutar Todos", command=self.ejecutar_todos,
                 bg="#e74c3c", fg="white", font=("Arial", 10, "bold"), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="⏹️ Detener", command=self.detener_todos,
                 bg="#7f8c8d", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="📈 Comparar Rendimiento", command=self.comparar_rendimiento,
                 bg="#3498db", fg="white", font=("Arial", 10), padx=15, pady=5, width=18).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="💾 Guardar Todo", command=self.guardar_todos_resultados,
                 bg="#2ecc71", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="🔄 Limpiar Todo", command=self.limpiar_todo,
                 bg="#95a5a6", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        
        # Mostrar datos actuales
        self.label_datos = tk.Label(datos_frame, text="Sin datos", bg="#1a1a2e", fg="#ecf0f1",
                                    font=("Arial", 10), wraplength=1200)
        self.label_datos.pack(pady=5)
        
        # Barra de estado
        self.status_label = tk.Label(top_panel, text="✅ Listo", bg="#1a1a2e", fg="#2ecc71",
                                     font=("Arial", 9))
        self.status_label.pack(pady=5)
        
        # ========== PANEL INFERIOR ==========
        # Frame para scroll de métodos
        canvas_metodos = tk.Canvas(bottom_panel, bg="#1a1a2e", highlightthickness=0)
        scrollbar_vertical = ttk.Scrollbar(bottom_panel, orient="vertical", command=canvas_metodos.yview)
        scrollbar_horizontal = ttk.Scrollbar(bottom_panel, orient="horizontal", command=canvas_metodos.xview)
        
        self.frame_metodos = ttk.Frame(canvas_metodos)
        
        self.frame_metodos.bind("<Configure>", lambda e: canvas_metodos.configure(scrollregion=canvas_metodos.bbox("all")))
        canvas_metodos.create_window((0, 0), window=self.frame_metodos, anchor="nw")
        canvas_metodos.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)
        
        canvas_metodos.pack(side="left", fill="both", expand=True)
        scrollbar_vertical.pack(side="right", fill="y")
        scrollbar_horizontal.pack(side="bottom", fill="x")
        
        # Crear frames para cada método en grid
        self.frames_metodos = {}
        
        # Configurar grid con 4 columnas para mejor distribución
        for i in range(4):
            self.frame_metodos.grid_columnconfigure(i, weight=1)
        
        # Título de métodos internos
        interno_title = tk.Label(self.frame_metodos, text="🔹 MÉTODOS INTERNOS 🔹", 
                                 font=("Arial", 14, "bold"), bg="#1a1a2e", fg="#00d4ff")
        interno_title.grid(row=0, column=0, columnspan=4, pady=15)
        
        # Métodos Internos - organizados en 2 filas de 4 columnas
        metodos_internos_lista = list(self.metodos_internos.keys())
        
        # Fila 1 - primeros 4 métodos
        row = 1
        for col, nombre in enumerate(metodos_internos_lista[:4]):
            self._crear_frame_metodo_grid(nombre, row, col)
        
        # Fila 2 - siguientes métodos
        row = 2
        for col, nombre in enumerate(metodos_internos_lista[4:]):
            self._crear_frame_metodo_grid(nombre, row, col)
        
        # Título de métodos externos
        externo_title = tk.Label(self.frame_metodos, text="🔸 MÉTODOS EXTERNOS 🔸", 
                                 font=("Arial", 14, "bold"), bg="#1a1a2e", fg="#00d4ff")
        externo_title.grid(row=3, column=0, columnspan=4, pady=15)
        
        # Métodos Externos - centrados en 3 columnas
        row = 4
        metodos_externos_lista = list(self.metodos_externos.keys())
        
        # Calcular offset para centrar (3 métodos, centrarlos en 4 columnas)
        offset = (4 - len(metodos_externos_lista)) // 2
        
        for idx, nombre in enumerate(metodos_externos_lista):
            col = offset + idx
            self._crear_frame_metodo_grid(nombre, row, col)
    
    def _crear_frame_metodo_grid(self, nombre, row, col):
        """Crea un frame para un método de ordenamiento en grid"""
        frame = tk.Frame(self.frame_metodos, bg="#16213e", relief=tk.RAISED, bd=2, padx=8, pady=8)
        frame.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        
        # Título del método
        titulo_color = "#e74c3c" if nombre == "Quicksort" else "#f39c12" if nombre == "Radixsort" else "#00d4ff"
        
        label = tk.Label(frame, text=nombre, font=("Arial", 11, "bold"),
                        bg="#16213e", fg=titulo_color)
        label.pack(pady=5)
        
        # Canvas para visualización
        canvas = tk.Canvas(frame, bg="#0f3460", height=140, width=300)
        canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Frame para botones
        botones_frame = tk.Frame(frame, bg="#16213e")
        botones_frame.pack(fill=tk.X, pady=5)
        
        # Label para tiempo
        tiempo_label = tk.Label(frame, text="-- s", font=("Arial", 9),
                               bg="#16213e", fg="#f39c12")
        tiempo_label.pack(pady=2)
        
        # Botón ejecutar individual
        btn_ejecutar = tk.Button(botones_frame, text="▶️ Ejecutar", 
                                command=lambda n=nombre: self.ejecutar_individual(n),
                                bg="#2c3e50", fg="white", font=("Arial", 9), padx=12, pady=3)
        btn_ejecutar.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        # Botón guardar resultado individual
        btn_guardar = tk.Button(botones_frame, text="💾 Guardar", 
                               command=lambda n=nombre: self.guardar_resultado_individual(n),
                               bg="#27ae60", fg="white", font=("Arial", 9), padx=12, pady=3)
        btn_guardar.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        btn_guardar.config(state=tk.DISABLED)
        
        # Label para resultados texto (cuando hay muchos datos)
        texto_resultado = tk.Label(frame, text="", bg="#16213e", fg="#ecf0f1",
                                   font=("Arial", 8), wraplength=280, justify=tk.LEFT)
        
        self.frames_metodos[nombre] = {
            'frame': frame,
            'canvas': canvas,
            'tiempo_label': tiempo_label,
            'btn_ejecutar': btn_ejecutar,
            'btn_guardar': btn_guardar,
            'texto_resultado': texto_resultado,
            'ultimo_tiempo': None,
            'ultimo_resultado': None
        }
    
    def dibujar_barras(self, canvas, datos):
        """Dibuja gráfico de barras en un canvas"""
        canvas.delete("all")
        
        if not datos:
            return
        
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        
        if ancho < 10:
            ancho = 300
        if alto < 10:
            alto = 140
        
        n = len(datos)
        if n == 0:
            return
        
        ancho_barra = max(6, (ancho - 30) / n - 2)
        max_valor = max(datos) if datos else 100
        
        for i, valor in enumerate(datos):
            x0 = 15 + i * (ancho_barra + 2)
            y0 = alto - 25 - (valor / max_valor) * (alto - 50)
            x1 = x0 + ancho_barra
            y1 = alto - 25
            
            # Color basado en el valor
            r = int((valor / max_valor) * 255)
            b = int((1 - valor / max_valor) * 255)
            color = f'#{r:02x}40{b:02x}'
            canvas.create_rectangle(x0, y0, x1, y1, fill=color, outline="white", width=1)
            
            if ancho_barra > 12 and n <= 15:
                canvas.create_text(x0 + ancho_barra/2, y0 - 5, text=str(valor),
                                  font=("Arial", 8), fill="white")
    
    def ejecutar_individual(self, nombre):
        """Ejecuta un método específico"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        # Obtener el método
        metodo = None
        if nombre in self.metodos_internos:
            metodo = self.metodos_internos[nombre]
        elif nombre in self.metodos_externos:
            metodo = self.metodos_externos[nombre]
        
        if not metodo:
            return
        
        # Deshabilitar botones mientras ejecuta
        frame_info = self.frames_metodos[nombre]
        frame_info['btn_ejecutar'].config(state=tk.DISABLED, text="Ejecutando...")
        frame_info['btn_guardar'].config(state=tk.DISABLED)
        frame_info['tiempo_label'].config(text="... s")
        
        def ejecutar():
            datos_copy = self.datos_actuales.copy()
            
            start = time.time()
            
            # Ejecutar el método
            _, resultado = metodo(datos_copy)
            tiempo = time.time() - start
            
            # Guardar resultado
            frame_info['ultimo_tiempo'] = tiempo
            frame_info['ultimo_resultado'] = resultado
            
            # Visualizar resultado
            if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
                self.dibujar_barras(frame_info['canvas'], resultado)
                frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
                if frame_info['texto_resultado'].winfo_ismapped():
                    frame_info['texto_resultado'].pack_forget()
                    frame_info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            else:
                self.mostrar_resultado_texto(frame_info, resultado, tiempo)
            
            frame_info['btn_ejecutar'].config(state=tk.NORMAL, text="▶️ Ejecutar")
            frame_info['btn_guardar'].config(state=tk.NORMAL)
        
        thread = threading.Thread(target=ejecutar)
        thread.daemon = True
        thread.start()
    
    def mostrar_resultado_texto(self, frame_info, datos, tiempo):
        """Muestra resultados en modo texto cuando hay muchos datos"""
        frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
        frame_info['ultimo_tiempo'] = tiempo
        frame_info['ultimo_resultado'] = datos
        frame_info['btn_guardar'].config(state=tk.NORMAL)
        
        if len(datos) <= 50:
            texto = f"Resultado: {datos}"
        else:
            texto = f"Resultado ({len(datos)}):\n{datos[:20]}...\n...{datos[-20:]}"
        
        frame_info['texto_resultado'].config(text=texto, font=("Arial", 8))
        
        if not frame_info['texto_resultado'].winfo_ismapped():
            frame_info['canvas'].pack_forget()
            frame_info['texto_resultado'].pack(pady=5, padx=5, fill=tk.BOTH, expand=True)
    
    def ejecutar_todos(self):
        """Ejecuta todos los métodos simultáneamente"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        self.status_label.config(text="🔄 Ejecutando todos los métodos...")
        
        def ejecutar():
            threads = []
            
            for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
                frame_info = self.frames_metodos[nombre]
                frame_info['btn_ejecutar'].config(state=tk.DISABLED, text="Ejecutando...")
                frame_info['btn_guardar'].config(state=tk.DISABLED)
                frame_info['tiempo_label'].config(text="... s")
                
                metodo = None
                if nombre in self.metodos_internos:
                    metodo = self.metodos_internos[nombre]
                else:
                    metodo = self.metodos_externos[nombre]
                
                thread = threading.Thread(target=self._ejecutar_metodo_hilo, 
                                          args=(nombre, metodo, frame_info))
                thread.daemon = True
                thread.start()
                threads.append(thread)
            
            for thread in threads:
                thread.join()
            
            self.status_label.config(text="✅ Todos los métodos completados")
        
        threading.Thread(target=ejecutar, daemon=True).start()
    
    def _ejecutar_metodo_hilo(self, nombre, metodo, frame_info):
        """Ejecuta un método en un hilo"""
        datos_copy = self.datos_actuales.copy()
        
        start = time.time()
        _, resultado = metodo(datos_copy)
        tiempo = time.time() - start
        
        frame_info['ultimo_tiempo'] = tiempo
        frame_info['ultimo_resultado'] = resultado
        
        if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
            self.dibujar_barras(frame_info['canvas'], resultado)
            frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
            if frame_info['texto_resultado'].winfo_ismapped():
                frame_info['texto_resultado'].pack_forget()
                frame_info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        else:
            self.mostrar_resultado_texto(frame_info, resultado, tiempo)
        
        frame_info['btn_ejecutar'].config(state=tk.NORMAL, text="▶️ Ejecutar")
        frame_info['btn_guardar'].config(state=tk.NORMAL)
    
    # Los siguientes métodos (generar_aleatorios, ingresar_manual, cargar_archivo_txt,
    # cargar_archivo_excel, cargar_archivo_json, guardar_resultado_individual,
    # comparar_rendimiento, guardar_todos_resultados, limpiar_todo, detener_todos)
    # se mantienen igual que en tu código original...
    
    def generar_aleatorios(self):
        """Genera datos aleatorios"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Generar Datos")
        dialog.geometry("350x180")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        tk.Label(dialog, text="Número de elementos:", bg="#1a1a2e", fg="white", 
                font=("Arial", 10)).pack(pady=10)
        
        var = tk.IntVar(value=15)
        entry = tk.Entry(dialog, textvariable=var, font=("Arial", 10), width=10)
        entry.pack(pady=5)
        
        def aceptar():
            try:
                n = var.get()
                if n <= 0:
                    messagebox.showerror("Error", "Ingrese un número positivo")
                    return
                
                if n > 1000:
                    if not messagebox.askyesno("Advertencia", f"¿Generar {n} números? Puede ser lento."):
                        return
                
                self.datos_actuales = [random.randint(1, 1000) for _ in range(n)]
                
                # Limpiar resultados anteriores
                self._limpiar_resultados_metodos()
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.nombre_archivo_cargado = ""
                self.status_label.config(text=f"✅ Generados {len(self.datos_actuales)} datos aleatorios")
                
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                dialog.destroy()
            except:
                messagebox.showerror("Error", "Ingrese un número válido")
        
        tk.Button(dialog, text="Aceptar", command=aceptar, bg="#27ae60", fg="white", 
                 font=("Arial", 10), padx=20, pady=5).pack(pady=10)
    
    def _limpiar_resultados_metodos(self):
        """Limpia los resultados almacenados de los métodos"""
        for nombre, info in self.frames_metodos.items():
            info['ultimo_resultado'] = None
            info['ultimo_tiempo'] = None
            info['btn_guardar'].config(state=tk.DISABLED)
            info['canvas'].delete("all")
            info['tiempo_label'].config(text="-- s")
            if info['texto_resultado'].winfo_ismapped():
                info['texto_resultado'].pack_forget()
            info['texto_resultado'].config(text="")
            if not info['canvas'].winfo_ismapped():
                info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def ingresar_manual(self):
        """Ingreso manual de datos"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Ingresar Datos")
        dialog.geometry("550x350")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        tk.Label(dialog, text="Ingrese números separados por espacios o comas:", 
                bg="#1a1a2e", fg="white", font=("Arial", 10)).pack(pady=10)
        
        text_area = tk.Text(dialog, height=10, width=60, font=("Arial", 10), bg="#0f3460", fg="white")
        text_area.pack(pady=10, padx=10)
        
        tk.Label(dialog, text="Ejemplo: 15 8 23 42 16 4", 
                bg="#1a1a2e", fg="#7f8c8d", font=("Arial", 9)).pack()
        
        def aceptar():
            texto = text_area.get("1.0", tk.END)
            texto = texto.replace(',', ' ')
            numeros = texto.split()
            try:
                datos = [int(n) for n in numeros if n.strip().isdigit()]
                if not datos:
                    messagebox.showerror("Error", "Ingrese al menos un número")
                    return
                
                if len(datos) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se ingresaron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos
                
                self._limpiar_resultados_metodos()
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.nombre_archivo_cargado = ""
                self.status_label.config(text=f"✅ Ingresados {len(datos)} datos manualmente")
                
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                dialog.destroy()
            except:
                messagebox.showerror("Error", "Ingrese solo números válidos")
        
        btn_frame = tk.Frame(dialog, bg="#1a1a2e")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Aceptar", command=aceptar, bg="#27ae60", fg="white", padx=20).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancelar", command=dialog.destroy, bg="#e74c3c", fg="white", padx=20).pack(side=tk.LEFT, padx=5)
    
    def cargar_archivo_txt(self):
        """Carga datos desde archivo de texto"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo TXT",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    contenido = f.read()
                
                numeros = re.findall(r'-?\d+', contenido)
                datos = [int(n) for n in numeros]
                
                if datos:
                    if len(datos) > 1000:
                        if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                            return
                    
                    self.datos_actuales = datos
                    self._limpiar_resultados_metodos()
                    self.nombre_archivo_cargado = archivo
                    
                    if len(self.datos_actuales) <= 30:
                        self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                    else:
                        self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                    
                    self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                    self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                else:
                    messagebox.showerror("Error", "No se encontraron números en el archivo")
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo: {str(e)}")
    
    def cargar_archivo_json(self):
        """Carga datos desde archivo JSON"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo JSON",
            filetypes=[("Archivos JSON", "*.json"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                datos = []
                
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, (int, float)):
                            datos.append(int(item))
                elif isinstance(data, dict):
                    for key, value in data.items():
                        if isinstance(value, list):
                            for item in value:
                                if isinstance(item, (int, float)):
                                    datos.append(int(item))
                        elif isinstance(value, (int, float)):
                            datos.append(int(value))
                
                if not datos:
                    messagebox.showerror("Error", "No se encontraron números en el archivo JSON")
                    return
                
                if len(datos) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos
                self._limpiar_resultados_metodos()
                self.nombre_archivo_cargado = archivo
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos JSON ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos JSON ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo: {str(e)}")
    
    def cargar_archivo_excel(self):
        """Carga datos desde archivo Excel (.xlsx)"""
        if not EXCEL_DISPONIBLE:
            messagebox.showerror("Error", "openpyxl no está instalado.\n\nInstálalo con: pip install openpyxl")
            return
        
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                wb = load_workbook(archivo, data_only=True)
                self._seleccionar_hoja_excel(archivo, wb, wb.sheetnames)
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo Excel:\n{str(e)}")
    
    def columna_a_indice(self, columna):
        """Convierte letras de columna a índice numérico"""
        columna = columna.upper().strip()
        indice = 0
        for char in columna:
            if 'A' <= char <= 'Z':
                indice = indice * 26 + (ord(char) - ord('A') + 1)
        return indice
    
    def _seleccionar_hoja_excel(self, archivo, wb, hojas):
        """Diálogo para seleccionar hoja y columna del Excel"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Seleccionar datos del Excel")
        dialog.geometry("600x500")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        main_frame = tk.Frame(dialog, bg="#1a1a2e")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Selección de hoja
        tk.Label(main_frame, text="Seleccione la hoja:", bg="#1a1a2e", fg="white").pack(pady=5)
        
        hoja_var = tk.StringVar(value=hojas[0] if hojas else "")
        hoja_combo = ttk.Combobox(main_frame, textvariable=hoja_var, values=hojas, state="readonly")
        hoja_combo.pack(pady=5)
        
        # Selección de columna
        tk.Label(main_frame, text="Seleccione la columna (ej: A, B, C):", bg="#1a1a2e", fg="white").pack(pady=5)
        columna_var = tk.StringVar(value="A")
        columna_entry = tk.Entry(main_frame, textvariable=columna_var, width=10)
        columna_entry.pack(pady=5)
        
        # Opción de encabezado
        incluir_encabezado = tk.BooleanVar(value=False)
        tk.Checkbutton(main_frame, text="La primera fila es encabezado (ignorar)",
                      variable=incluir_encabezado, bg="#1a1a2e", fg="white",
                      selectcolor="#1a1a2e").pack(pady=5)
        
        def aceptar():
            try:
                hoja = hoja_var.get()
                columna = columna_var.get().strip()
                ws = wb[hoja]
                
                col_idx = self.columna_a_indice(columna)
                
                datos = []
                start_row = 2 if incluir_encabezado.get() else 1
                
                for row in range(start_row, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx).value
                    if cell is not None:
                        try:
                            datos.append(int(float(cell)))
                        except (ValueError, TypeError):
                            pass
                
                if not datos:
                    messagebox.showerror("Error", "No se encontraron números en la columna seleccionada")
                    return
                
                if len(datos) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos
                self._limpiar_resultados_metodos()
                self.nombre_archivo_cargado = archivo
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos Excel ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos Excel ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar Excel:\n{str(e)}")
        
        btn_frame = tk.Frame(dialog, bg="#1a1a2e")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Aceptar", command=aceptar, bg="#27ae60", fg="white",
                 padx=20, pady=5).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="Cancelar", command=dialog.destroy, bg="#e74c3c", fg="white",
                 padx=20, pady=5).pack(side=tk.LEFT, padx=10)
    
    def guardar_resultado_individual(self, nombre):
        """Guarda solo el resultado de un método específico"""
        frame_info = self.frames_metodos[nombre]
        
        if frame_info['ultimo_resultado'] is None:
            messagebox.showwarning("Advertencia", f"Primero ejecute el método '{nombre}'")
            return
        
        from tkinter import filedialog
        
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo_default = f"{nombre}_{len(self.datos_actuales)}elem_{fecha_hora}.txt"
        
        archivo = filedialog.asksaveasfilename(
            title=f"Guardar resultado de {nombre}",
            defaultextension=".txt",
            initialfile=nombre_archivo_default,
            filetypes=[
                ("Archivos de texto", "*.txt"),
                ("Archivos CSV", "*.csv"),
                ("Archivos JSON", "*.json")
            ]
        )
        
        if archivo:
            try:
                with open(archivo, 'w', encoding='utf-8') as f:
                    f.write("=" * 60 + "\n")
                    f.write(f"RESULTADO DE ORDENAMIENTO - {nombre}\n")
                    f.write("=" * 60 + "\n\n")
                    f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Cantidad de elementos: {len(self.datos_actuales)}\n")
                    f.write(f"Tiempo: {frame_info['ultimo_tiempo']:.6f} segundos\n\n")
                    f.write("Datos originales:\n")
                    f.write(f"{self.datos_actuales}\n\n")
                    f.write("Resultado ordenado:\n")
                    f.write(f"{frame_info['ultimo_resultado']}\n")
                
                self.status_label.config(text=f"✅ Resultado de {nombre} guardado")
                messagebox.showinfo("Éxito", f"Resultado de '{nombre}' guardado correctamente")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def comparar_rendimiento(self):
        """Compara el rendimiento de todos los métodos"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        resultados = {}
        
        # Medir tiempos para cada método
        for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
            metodo = None
            if nombre in self.metodos_internos:
                metodo = self.metodos_internos[nombre]
            else:
                metodo = self.metodos_externos[nombre]
            
            datos_copy = self.datos_actuales.copy()
            start = time.time()
            _, _ = metodo(datos_copy)
            tiempo = time.time() - start
            resultados[nombre] = tiempo
        
        # Mostrar resultados
        ventana_resultados = tk.Toplevel(self.ventana)
        ventana_resultados.title("Comparación de Rendimiento")
        ventana_resultados.geometry("550x500")
        ventana_resultados.configure(bg="#1a1a2e")
        
        text_area = tk.Text(ventana_resultados, bg="#0a0a1a", fg="#00ff00", font=("Courier", 10))
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_area.insert(tk.END, "=" * 60 + "\n")
        text_area.insert(tk.END, "COMPARACIÓN DE RENDIMIENTO\n")
        text_area.insert(tk.END, "=" * 60 + "\n\n")
        text_area.insert(tk.END, f"Elementos: {len(self.datos_actuales)}\n\n")
        text_area.insert(tk.END, f"{'Método':<20} {'Tiempo (segundos)':<20}\n")
        text_area.insert(tk.END, "-" * 60 + "\n")
        
        for nombre, tiempo in sorted(resultados.items(), key=lambda x: x[1]):
            text_area.insert(tk.END, f"{nombre:<20} {tiempo:.6f}\n")
        
        text_area.insert(tk.END, "\n" + "=" * 60 + "\n")
        text_area.config(state=tk.DISABLED)
        
        btn_frame = tk.Frame(ventana_resultados, bg="#1a1a2e")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Cerrar", command=ventana_resultados.destroy,
                 bg="#e74c3c", fg="white", padx=20, pady=5).pack()
    
    def guardar_todos_resultados(self):
        """Guarda los resultados de todos los métodos ejecutados"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "No hay datos para guardar")
            return
        
        resultados_existentes = {nombre: info for nombre, info in self.frames_metodos.items() 
                                if info['ultimo_resultado'] is not None}
        
        if not resultados_existentes:
            messagebox.showwarning("Advertencia", "No hay resultados para guardar")
            return
        
        from tkinter import filedialog
        
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo_default = f"todos_resultados_{len(self.datos_actuales)}elem_{fecha_hora}.txt"
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar todos los resultados",
            defaultextension=".txt",
            initialfile=nombre_archivo_default,
            filetypes=[("Archivos de texto", "*.txt"), ("Archivos JSON", "*.json")]
        )
        
        if archivo:
            try:
                with open(archivo, 'w', encoding='utf-8') as f:
                    f.write("=" * 80 + "\n")
                    f.write("REPORTE COMPLETO DE ORDENAMIENTO\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Cantidad de elementos: {len(self.datos_actuales)}\n\n")
                    f.write("DATOS ORIGINALES:\n")
                    f.write(f"{self.datos_actuales}\n\n")
                    f.write("-" * 80 + "\n\n")
                    
                    for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
                        if nombre in resultados_existentes:
                            info = resultados_existentes[nombre]
                            f.write(f"🔸 {nombre}:\n")
                            f.write(f"   Tiempo: {info['ultimo_tiempo']:.6f} segundos\n")
                            f.write(f"   Resultado: {info['ultimo_resultado']}\n\n")
                
                self.status_label.config(text=f"✅ Todos los resultados guardados")
                messagebox.showinfo("Éxito", "Todos los resultados guardados correctamente")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def limpiar_todo(self):
        """Limpia todos los datos y resultados"""
        if messagebox.askyesno("Confirmar", "¿Está seguro de limpiar todos los datos y resultados?"):
            self.datos_actuales = []
            self.label_datos.config(text="Sin datos")
            self.nombre_archivo_cargado = ""
            self._limpiar_resultados_metodos()
            self.status_label.config(text="✅ Todo limpiado")
    
    def detener_todos(self):
        """Detiene la ejecución"""
        self.status_label.config(text="⏹️ Detenido")
        messagebox.showinfo("Info", "Las ejecuciones en curso continuarán hasta finalizar")
    
    def ejecutar(self):
        """Inicia la aplicación"""
        self.ventana.mainloop()


if __name__ == "__main__":
    # Verificar que las librerías existen
    try:
        import ordenamiento_interno
        import ordenamiento_externo
        print("✅ Librerías de ordenamiento cargadas correctamente")
    except ImportError as e:
        print(f"❌ Error al cargar librerías: {e}")
        print("Asegúrate de que los archivos 'ordenamiento_interno.py' y 'ordenamiento_externo.py' estén en el mismo directorio")
        exit(1)
    
    app = DemoCompletaOrdenamiento()
    app.ejecutar()