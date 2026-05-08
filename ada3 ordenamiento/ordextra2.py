import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import time
import threading
import random
import math
import os
import json
import csv
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
import hashlib
import re
from collections import Counter

# Importar librerías adicionales (requieren instalación)
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image
    import pytesseract
    IMAGE_OCR_AVAILABLE = True
except ImportError:
    IMAGE_OCR_AVAILABLE = False

try:
    import zipfile
    import tarfile
    COMPRESS_AVAILABLE = True
except ImportError:
    COMPRESS_AVAILABLE = True

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

class OrdenamientoExternoSimultaneo:
    def __init__(self, root, datos_usuario=None, archivo_cargado=None):
        self.root = root
        self.root.title("Ordenamiento Externo - Comparación de Dos Archivos")
        self.root.geometry("1400x900")
        self.root.configure(bg="#2c3e50")
        
        # Tipo de dato a analizar
        self.tipo_dato = "numeros"  # Opciones: numeros, palabras, lineas, fechas, todo
        self.comparador = self.comparar_numeros  # Función de comparación por defecto
        
        # Datos de los dos archivos
        self.datos_archivo1 = []
        self.datos_archivo2 = []
        self.datos_raw_archivo1 = []  # Datos sin procesar para mostrar
        self.datos_raw_archivo2 = []
        self.nombre_archivo1 = None
        self.nombre_archivo2 = None
        self.ruta_archivo1 = None
        self.ruta_archivo2 = None
        
        # Configuración de hojas de Excel
        self.hojas_archivo1 = []  # Lista de hojas seleccionadas para archivo 1
        self.hojas_archivo2 = []  # Lista de hojas seleccionadas para archivo 2
        self.hojas_disponibles1 = []  # Todas las hojas disponibles en archivo 1
        self.hojas_disponibles2 = []  # Todas las hojas disponibles en archivo 2
        
        # Datos combinados y resultados
        self.datos_combinados = []
        self.datos_ejemplo = []
        self.resultados = {}
        self.resultados_comparacion = {}
        
        # Crear canvas principal con scrollbar
        self.main_canvas = tk.Canvas(self.root, bg="#2c3e50", highlightthickness=0)
        self.main_scrollbar = tk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)
        
        self.main_scrollbar.pack(side="right", fill="y")
        self.main_canvas.pack(side="left", fill="both", expand=True)
        
        # Frame principal que contendrá todo (dentro del canvas)
        self.main_frame = tk.Frame(self.main_canvas, bg="#2c3e50")
        self.main_canvas_window = self.main_canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        
        # Configurar el canvas para que se ajuste al frame
        self.main_frame.bind("<Configure>", self.on_frame_configure)
        self.main_canvas.bind("<Configure>", self.on_canvas_configure)
        
        # Permitir scroll con la rueda del mouse
        self.main_canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        
        self.crear_widgets()
    
    def on_frame_configure(self, event=None):
        """Actualizar el tamaño del canvas cuando cambia el frame"""
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def on_canvas_configure(self, event):
        """Ajustar el ancho del frame cuando cambia el canvas"""
        self.main_canvas.itemconfig(self.main_canvas_window, width=event.width)
    
    def on_mousewheel(self, event):
        """Manejar el scroll con la rueda del mouse"""
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def crear_widgets(self):
        # Título principal
        titulo = tk.Label(self.main_frame, text="COMPARACIÓN DE DOS ARCHIVOS - ORDENAMIENTO EXTERNO", 
                         font=("Arial", 16, "bold"), bg="#2c3e50", fg="white", pady=10)
        titulo.pack(fill=tk.X)
        
        # Frame para selección de tipo de dato
        frame_tipo_dato = tk.LabelFrame(self.main_frame, text="TIPO DE DATO A ANALIZAR", 
                                        font=("Arial", 10, "bold"), bg="#34495e", fg="white", padx=10, pady=5)
        frame_tipo_dato.pack(fill=tk.X, padx=10, pady=5)
        
        self.tipo_dato_var = tk.StringVar(value="numeros")
        
        tipos = [
            ("🔢 Números", "numeros"),
            ("📝 Palabras", "palabras"),
            ("📄 Líneas completas", "lineas"),
            ("📅 Fechas", "fechas"),
            ("🔤 Caracteres", "caracteres"),
            ("🌐 Todo (texto completo)", "todo")
        ]
        
        for i, (texto, valor) in enumerate(tipos):
            rb = tk.Radiobutton(frame_tipo_dato, text=texto, variable=self.tipo_dato_var, value=valor,
                               bg="#34495e", fg="white", selectcolor="#2c3e50",
                               font=("Arial", 9), command=self.cambiar_tipo_dato)
            rb.pack(side=tk.LEFT, padx=10, pady=5)
        
        # Frame para los dos archivos
        frame_archivos = tk.Frame(self.main_frame, bg="#34495e", pady=10)
        frame_archivos.pack(fill=tk.X, padx=10, pady=5)
        
        # Archivo 1
        frame_archivo1 = tk.LabelFrame(frame_archivos, text="ARCHIVO 1", font=("Arial", 10, "bold"),
                                       bg="#ecf0f1", fg="#3498db", padx=10, pady=5)
        frame_archivo1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        btn_cargar1 = tk.Button(frame_archivo1, text="📂 CARGAR ARCHIVO 1", command=lambda: self.cargar_archivo(1),
                               bg="#3498db", fg="white", font=("Arial", 10, "bold"),
                               width=20, cursor="hand2")
        btn_cargar1.pack(pady=5)
        
        self.lbl_archivo1 = tk.Label(frame_archivo1, text="Ningún archivo cargado", 
                                     bg="#ecf0f1", fg="#7f8c8d", font=("Arial", 9))
        self.lbl_archivo1.pack(pady=5)
        
        self.lbl_hojas1 = tk.Label(frame_archivo1, text="", bg="#ecf0f1", fg="#3498db", 
                                   font=("Arial", 8, "italic"))
        self.lbl_hojas1.pack(pady=2)
        
        self.lbl_stats1 = tk.Label(frame_archivo1, text="", bg="#ecf0f1", fg="#2c3e50", font=("Arial", 9))
        self.lbl_stats1.pack(pady=5)
        
        # Archivo 2
        frame_archivo2 = tk.LabelFrame(frame_archivos, text="ARCHIVO 2", font=("Arial", 10, "bold"),
                                       bg="#ecf0f1", fg="#e74c3c", padx=10, pady=5)
        frame_archivo2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        btn_cargar2 = tk.Button(frame_archivo2, text="📂 CARGAR ARCHIVO 2", command=lambda: self.cargar_archivo(2),
                               bg="#e74c3c", fg="white", font=("Arial", 10, "bold"),
                               width=20, cursor="hand2")
        btn_cargar2.pack(pady=5)
        
        self.lbl_archivo2 = tk.Label(frame_archivo2, text="Ningún archivo cargado", 
                                     bg="#ecf0f1", fg="#7f8c8d", font=("Arial", 9))
        self.lbl_archivo2.pack(pady=5)
        
        self.lbl_hojas2 = tk.Label(frame_archivo2, text="", bg="#ecf0f1", fg="#e74c3c", 
                                   font=("Arial", 8, "italic"))
        self.lbl_hojas2.pack(pady=2)
        
        self.lbl_stats2 = tk.Label(frame_archivo2, text="", bg="#ecf0f1", fg="#2c3e50", font=("Arial", 9))
        self.lbl_stats2.pack(pady=5)
        
        # Botones de acción
        frame_botones = tk.Frame(self.main_frame, bg="#2c3e50", pady=10)
        frame_botones.pack(fill=tk.X)
        
        btn_comparar = tk.Button(frame_botones, text="🔍 COMPARAR ARCHIVOS", command=self.comparar_archivos,
                                bg="#9b59b6", fg="white", font=("Arial", 11, "bold"),
                                width=20, height=2, cursor="hand2")
        btn_comparar.pack(side=tk.LEFT, padx=10, expand=True)
        
        btn_combinar = tk.Button(frame_botones, text="🔄 COMBINAR Y ORDENAR", command=self.combinar_y_ordenar,
                                bg="#f39c12", fg="white", font=("Arial", 11, "bold"),
                                width=20, height=2, cursor="hand2")
        btn_combinar.pack(side=tk.LEFT, padx=10, expand=True)
        
        btn_guardar = tk.Button(frame_botones, text="💾 GUARDAR RESULTADO", command=self.guardar_resultado_combinado,
                               bg="#27ae60", fg="white", font=("Arial", 11, "bold"),
                               width=20, height=2, cursor="hand2")
        btn_guardar.pack(side=tk.LEFT, padx=10, expand=True)
        
        # Mostrar información de combinación
        self.lbl_combinacion = tk.Label(self.main_frame, text="", bg="#2c3e50", fg="#f39c12", 
                                        font=("Arial", 10, "bold"), pady=5)
        self.lbl_combinacion.pack(fill=tk.X)
        
        # Frame para visualización de datos (con scroll)
        frame_visual_container = tk.Frame(self.main_frame, bg="#2c3e50")
        frame_visual_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Título del área de visualización
        lbl_visual = tk.Label(frame_visual_container, text="VISTA PREVIA DE DATOS", 
                             font=("Arial", 10, "bold"), bg="#2c3e50", fg="white")
        lbl_visual.pack(anchor="w")
        
        # Canvas para mostrar los datos con scroll
        self.datos_canvas = tk.Canvas(frame_visual_container, bg="white", height=200, highlightthickness=2)
        datos_scrollbar = tk.Scrollbar(frame_visual_container, orient="vertical", command=self.datos_canvas.yview)
        self.datos_canvas.configure(yscrollcommand=datos_scrollbar.set)
        
        self.datos_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        datos_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Frame interno para los datos (permitirá scroll vertical)
        self.datos_inner_frame = tk.Frame(self.datos_canvas, bg="white")
        self.datos_canvas_window = self.datos_canvas.create_window((0, 0), window=self.datos_inner_frame, anchor="nw")
        
        self.datos_inner_frame.bind("<Configure>", self.on_datos_frame_configure)
        self.datos_canvas.bind("<Configure>", self.on_datos_canvas_configure)
        
        # Frame para los algoritmos (con scroll)
        algoritmos_container = tk.Frame(self.main_frame, bg="#2c3e50")
        algoritmos_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        lbl_algoritmos = tk.Label(algoritmos_container, text="ALGORITMOS DE ORDENAMIENTO", 
                                 font=("Arial", 10, "bold"), bg="#2c3e50", fg="white")
        lbl_algoritmos.pack(anchor="w")
        
        # Canvas con scroll para los algoritmos
        self.algoritmos_canvas = tk.Canvas(algoritmos_container, bg="#2c3e50", highlightthickness=0)
        algoritmos_scrollbar = tk.Scrollbar(algoritmos_container, orient="vertical", command=self.algoritmos_canvas.yview)
        self.algoritmos_canvas.configure(yscrollcommand=algoritmos_scrollbar.set)
        
        self.algoritmos_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        algoritmos_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Frame interno para los algoritmos
        self.algoritmos_inner_frame = tk.Frame(self.algoritmos_canvas, bg="#2c3e50")
        self.algoritmos_canvas_window = self.algoritmos_canvas.create_window((0, 0), window=self.algoritmos_inner_frame, anchor="nw")
        
        self.algoritmos_inner_frame.bind("<Configure>", self.on_algoritmos_frame_configure)
        self.algoritmos_canvas.bind("<Configure>", self.on_algoritmos_canvas_configure)
        
        # Configurar grid para los algoritmos
        self.algoritmos_inner_frame.columnconfigure(0, weight=1)
        self.algoritmos_inner_frame.columnconfigure(1, weight=1)
        self.algoritmos_inner_frame.columnconfigure(2, weight=1)
        
        self.frame_mezcla_directa = self.crear_panel_algoritmo(0, "MEZCLA DIRECTA", "#9b59b6")
        self.frame_mezcla_equilibrada = self.crear_panel_algoritmo(1, "MEZCLA EQUILIBRADA", "#e74c3c")
        self.frame_intercalacion = self.crear_panel_algoritmo(2, "INTERCALACIÓN", "#16a085")
        
        # Barra de progreso
        self.progress = ttk.Progressbar(self.main_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, padx=200, pady=5)
        
        self.ejecutando = False
    
    def cambiar_tipo_dato(self):
        """Cambiar el tipo de dato a analizar"""
        self.tipo_dato = self.tipo_dato_var.get()
        
        # Configurar el comparador según el tipo de dato
        if self.tipo_dato == "numeros":
            self.comparador = self.comparar_numeros
            messagebox.showinfo("Tipo de dato", "Modo: NÚMEROS\nSe extraerán y ordenarán valores numéricos.")
        elif self.tipo_dato == "palabras":
            self.comparador = self.comparar_strings
            messagebox.showinfo("Tipo de dato", "Modo: PALABRAS\nSe extraerán palabras y se ordenarán alfabéticamente.")
        elif self.tipo_dato == "lineas":
            self.comparador = self.comparar_strings
            messagebox.showinfo("Tipo de dato", "Modo: LÍNEAS\nSe ordenarán líneas completas alfabéticamente.")
        elif self.tipo_dato == "fechas":
            self.comparador = self.comparar_fechas
            messagebox.showinfo("Tipo de dato", "Modo: FECHAS\nSe detectarán y ordenarán fechas cronológicamente.")
        elif self.tipo_dato == "caracteres":
            self.comparador = self.comparar_strings
            messagebox.showinfo("Tipo de dato", "Modo: CARACTERES\nSe ordenarán caracteres individuales.")
        else:
            self.comparador = self.comparar_strings
            messagebox.showinfo("Tipo de dato", "Modo: TODO\nSe procesará el texto completo.")
        
        # Recargar archivos si están cargados
        if self.ruta_archivo1:
            self.recargar_archivo(1, self.ruta_archivo1)
        if self.ruta_archivo2:
            self.recargar_archivo(2, self.ruta_archivo2)
    
    def seleccionar_hojas_excel(self, num_archivo, ruta, hojas_disponibles):
        """Abrir ventana para seleccionar hojas de Excel"""
        ventana_hojas = tk.Toplevel(self.root)
        ventana_hojas.title(f"Seleccionar Hojas - Archivo {num_archivo}")
        ventana_hojas.geometry("500x400")
        ventana_hojas.configure(bg="#ecf0f1")
        
        tk.Label(ventana_hojas, text=f"SELECCIONAR HOJAS DE EXCEL", 
                font=("Arial", 12, "bold"), bg="#2c3e50", fg="white", pady=10).pack(fill=tk.X)
        
        tk.Label(ventana_hojas, text=f"Archivo: {os.path.basename(ruta)}", 
                font=("Arial", 10), bg="#ecf0f1", pady=5).pack()
        
        tk.Label(ventana_hojas, text="Seleccione las hojas que desea procesar:", 
                font=("Arial", 10), bg="#ecf0f1", pady=5).pack()
        
        # Frame para los checkboxes con scroll
        frame_scroll = tk.Frame(ventana_hojas, bg="#ecf0f1")
        frame_scroll.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        canvas = tk.Canvas(frame_scroll, bg="#ecf0f1", highlightthickness=0)
        scrollbar = tk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#ecf0f1")
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Variables para los checkboxes
        vars_hojas = {}
        for hoja in hojas_disponibles:
            var = tk.BooleanVar(value=True)  # Seleccionar todas por defecto
            vars_hojas[hoja] = var
            cb = tk.Checkbutton(scrollable_frame, text=hoja, variable=var, 
                               bg="#ecf0f1", font=("Arial", 10), anchor="w")
            cb.pack(fill=tk.X, padx=10, pady=2)
        
        # Botones de selección rápida
        frame_botones = tk.Frame(ventana_hojas, bg="#ecf0f1")
        frame_botones.pack(pady=10)
        
        def seleccionar_todas():
            for var in vars_hojas.values():
                var.set(True)
        
        def seleccionar_ninguna():
            for var in vars_hojas.values():
                var.set(False)
        
        tk.Button(frame_botones, text="Seleccionar Todas", command=seleccionar_todas,
                 bg="#3498db", fg="white", width=15).pack(side=tk.LEFT, padx=5)
        
        tk.Button(frame_botones, text="Seleccionar Ninguna", command=seleccionar_ninguna,
                 bg="#95a5a6", fg="white", width=15).pack(side=tk.LEFT, padx=5)
        
        def aceptar():
            hojas_seleccionadas = [hoja for hoja, var in vars_hojas.items() if var.get()]
            if not hojas_seleccionadas:
                messagebox.showwarning("Advertencia", "Debe seleccionar al menos una hoja.")
                return
            
            if num_archivo == 1:
                self.hojas_archivo1 = hojas_seleccionadas
                self.lbl_hojas1.config(text=f"Hojas: {len(hojas_seleccionadas)} seleccionadas")
            else:
                self.hojas_archivo2 = hojas_seleccionadas
                self.lbl_hojas2.config(text=f"Hojas: {len(hojas_seleccionadas)} seleccionadas")
            
            ventana_hojas.destroy()
            self.procesar_archivo_excel(num_archivo, ruta, hojas_seleccionadas)
        
        tk.Button(ventana_hojas, text="ACEPTAR", command=aceptar,
                 bg="#27ae60", fg="white", font=("Arial", 10, "bold"),
                 width=15, pady=5).pack(pady=10)
    
    def procesar_archivo_excel(self, num_archivo, ruta, hojas_seleccionadas):
        """Procesar el archivo Excel con las hojas seleccionadas"""
        try:
            wb = load_workbook(ruta, data_only=True)
            todos_los_datos = []
            contenido_raw = ""
            
            for nombre_hoja in hojas_seleccionadas:
                if nombre_hoja in wb.sheetnames:
                    ws = wb[nombre_hoja]
                    contenido_raw += f"\n--- HOJA: {nombre_hoja} ---\n"
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                valor = str(cell.value)
                                contenido_raw += valor + " "
                                todos_los_datos.append(valor)
                    contenido_raw += "\n"
            
            # Procesar según el tipo de dato
            datos_procesados = self.procesar_datos_segun_tipo(contenido_raw)
            
            if num_archivo == 1:
                self.datos_archivo1 = datos_procesados
                self.datos_raw_archivo1 = contenido_raw
                self.lbl_stats1.config(text=self.obtener_stats_texto(datos_procesados, contenido_raw))
            else:
                self.datos_archivo2 = datos_procesados
                self.datos_raw_archivo2 = contenido_raw
                self.lbl_stats2.config(text=self.obtener_stats_texto(datos_procesados, contenido_raw))
            
            self.mostrar_datos_en_canvas()
            
            messagebox.showinfo("Éxito", f"Archivo {num_archivo} procesado correctamente.\n"
                                        f"Hojas procesadas: {len(hojas_seleccionadas)}\n"
                                        f"Elementos extraídos: {len(datos_procesados)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el archivo Excel:\n{str(e)}")
    
    def recargar_archivo(self, num_archivo, ruta):
        """Recargar archivo con el nuevo tipo de dato"""
        try:
            extension = Path(ruta).suffix.lower()
            
            # Si es Excel, mantener las hojas seleccionadas
            if extension in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
                hojas_a_usar = self.hojas_archivo1 if num_archivo == 1 else self.hojas_archivo2
                if hojas_a_usar:
                    self.procesar_archivo_excel(num_archivo, ruta, hojas_a_usar)
                else:
                    # Si no hay hojas seleccionadas, mostrar selector
                    wb = load_workbook(ruta, data_only=True)
                    hojas_disponibles = wb.sheetnames
                    wb.close()
                    self.seleccionar_hojas_excel(num_archivo, ruta, hojas_disponibles)
            else:
                datos_raw = self.leer_archivo_raw(ruta)
                datos_procesados = self.procesar_datos_segun_tipo(datos_raw)
                
                if num_archivo == 1:
                    self.datos_archivo1 = datos_procesados
                    self.datos_raw_archivo1 = datos_raw
                    self.lbl_stats1.config(text=self.obtener_stats_texto(datos_procesados, datos_raw))
                else:
                    self.datos_archivo2 = datos_procesados
                    self.datos_raw_archivo2 = datos_raw
                    self.lbl_stats2.config(text=self.obtener_stats_texto(datos_procesados, datos_raw))
                
                self.mostrar_datos_en_canvas()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al recargar archivo {num_archivo}:\n{str(e)}")
    
    def obtener_stats_texto(self, datos_procesados, datos_raw):
        """Obtener estadísticas según el tipo de dato"""
        if self.tipo_dato == "numeros":
            if datos_procesados:
                return f"Elementos: {len(datos_procesados)}\nMín: {min(datos_procesados)} | Máx: {max(datos_procesados)}\nSuma: {sum(datos_procesados)}"
            else:
                return "No se encontraron números"
        elif self.tipo_dato == "palabras":
            return f"Palabras: {len(datos_procesados)}\nÚnicas: {len(set(datos_procesados))}\nMás larga: {max((len(p), p) for p in datos_procesados)[1] if datos_procesados else 'N/A'}"
        elif self.tipo_dato == "lineas":
            return f"Líneas: {len(datos_procesados)}\nCaracteres totales: {sum(len(str(l)) for l in datos_procesados)}"
        elif self.tipo_dato == "fechas":
            return f"Fechas encontradas: {len(datos_procesados)}\nÚnicas: {len(set(str(x) for x in datos_procesados))}"
        else:
            return f"Elementos: {len(datos_procesados)}"
    
    def comparar_numeros(self, a, b):
        """Comparar dos números"""
        return a - b
    
    def comparar_strings(self, a, b):
        """Comparar dos strings (orden alfabético)"""
        if a < b:
            return -1
        elif a > b:
            return 1
        return 0
    
    def comparar_fechas(self, a, b):
        """Comparar dos fechas"""
        if a < b:
            return -1
        elif a > b:
            return 1
        return 0
    
    def on_datos_frame_configure(self, event=None):
        """Actualizar el tamaño del canvas de datos cuando cambia el frame"""
        self.datos_canvas.configure(scrollregion=self.datos_canvas.bbox("all"))
    
    def on_datos_canvas_configure(self, event):
        """Ajustar el ancho del frame cuando cambia el canvas de datos"""
        self.datos_canvas.itemconfig(self.datos_canvas_window, width=event.width)
    
    def on_algoritmos_frame_configure(self, event=None):
        """Actualizar el tamaño del canvas de algoritmos cuando cambia el frame"""
        self.algoritmos_canvas.configure(scrollregion=self.algoritmos_canvas.bbox("all"))
    
    def on_algoritmos_canvas_configure(self, event):
        """Ajustar el ancho del frame cuando cambia el canvas de algoritmos"""
        self.algoritmos_canvas.itemconfig(self.algoritmos_canvas_window, width=event.width)
    
    def cargar_archivo(self, num_archivo):
        """Cargar archivo para comparación"""
        formatos = [
            ("Todos los archivos soportados", "*.txt *.csv *.json *.xml *.yaml *.yml *.log *.dat *.numbers *.xlsx *.xls *.pdf *.docx *.py *.java *.cpp *.c *.js *.html *.css *.sql *.zip *.tar *.gz"),
            ("📝 Texto", "*.txt *.log *.dat *.numbers"),
            ("📊 Datos", "*.csv *.json *.xml *.yaml *.yml"),
            ("📈 Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
        
        ruta = filedialog.askopenfilename(title=f"Seleccionar Archivo {num_archivo}", filetypes=formatos)
        
        if not ruta:
            return
        
        extension = Path(ruta).suffix.lower()
        
        # Si es archivo Excel, mostrar selector de hojas
        if extension in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
            try:
                wb = load_workbook(ruta, data_only=True)
                hojas_disponibles = wb.sheetnames
                wb.close()
                
                if num_archivo == 1:
                    self.hojas_disponibles1 = hojas_disponibles
                    self.ruta_archivo1 = ruta
                    self.nombre_archivo1 = os.path.basename(ruta)
                    self.lbl_archivo1.config(text=f"📄 {self.nombre_archivo1}")
                else:
                    self.hojas_disponibles2 = hojas_disponibles
                    self.ruta_archivo2 = ruta
                    self.nombre_archivo2 = os.path.basename(ruta)
                    self.lbl_archivo2.config(text=f"📄 {self.nombre_archivo2}")
                
                self.seleccionar_hojas_excel(num_archivo, ruta, hojas_disponibles)
                
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo leer el archivo Excel:\n{str(e)}")
        else:
            # Para otros tipos de archivo, procesar normalmente
            try:
                datos_raw = self.leer_archivo_raw(ruta)
                datos_procesados = self.procesar_datos_segun_tipo(datos_raw)
                
                if datos_procesados and len(datos_procesados) > 0:
                    if num_archivo == 1:
                        self.datos_archivo1 = datos_procesados
                        self.datos_raw_archivo1 = datos_raw
                        self.ruta_archivo1 = ruta
                        self.nombre_archivo1 = os.path.basename(ruta)
                        self.lbl_archivo1.config(text=f"📄 {self.nombre_archivo1}")
                        self.lbl_stats1.config(text=self.obtener_stats_texto(datos_procesados, datos_raw))
                        self.mostrar_datos_en_canvas()
                    else:
                        self.datos_archivo2 = datos_procesados
                        self.datos_raw_archivo2 = datos_raw
                        self.ruta_archivo2 = ruta
                        self.nombre_archivo2 = os.path.basename(ruta)
                        self.lbl_archivo2.config(text=f"📄 {self.nombre_archivo2}")
                        self.lbl_stats2.config(text=self.obtener_stats_texto(datos_procesados, datos_raw))
                        self.mostrar_datos_en_canvas()
                    
                    messagebox.showinfo("Éxito", f"Archivo {num_archivo} cargado correctamente.\n{len(datos_procesados)} elementos encontrados.\nTipo: {self.tipo_dato}")
                else:
                    messagebox.showwarning("Advertencia", f"No se encontraron elementos del tipo '{self.tipo_dato}' en el archivo.")
                    
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar el archivo:\n{str(e)}")
    
    def leer_archivo_raw(self, ruta):
        """Leer archivo y devolver contenido crudo según extensión"""
        extension = Path(ruta).suffix.lower()
        contenido = ""
        
        try:
            if extension in ['.txt', '.log', '.dat', '.numbers', '.ini', '.cfg', '.py', '.java', '.cpp', '.c', '.js', '.html', '.css', '.sql']:
                with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                    contenido = f.read()
            
            elif extension == '.csv':
                with open(ruta, 'r', encoding='utf-8') as f:
                    lector = csv.reader(f)
                    for fila in lector:
                        contenido += " ".join(fila) + "\n"
            
            elif extension == '.json':
                with open(ruta, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    contenido = json.dumps(data, ensure_ascii=False)
            
            elif extension == '.xml':
                tree = ET.parse(ruta)
                root = tree.getroot()
                contenido = ET.tostring(root, encoding='unicode')
            
            elif extension in ['.yaml', '.yml'] and YAML_AVAILABLE:
                with open(ruta, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    contenido = str(data)
            
            else:
                with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                    contenido = f.read()
            
            return contenido
            
        except Exception as e:
            raise Exception(f"Error leyendo archivo: {str(e)}")
    
    def procesar_datos_segun_tipo(self, contenido):
        """Procesar los datos según el tipo seleccionado"""
        if self.tipo_dato == "numeros":
            # Extraer números enteros y flotantes
            numeros = re.findall(r'-?\d+(?:\.\d+)?', contenido)
            # Convertir a int o float según corresponda
            resultado = []
            for n in numeros:
                try:
                    if '.' in n:
                        resultado.append(float(n))
                    else:
                        resultado.append(int(n))
                except:
                    pass
            return resultado
        
        elif self.tipo_dato == "palabras":
            # Extraer palabras (letras y números, mínimo 2 caracteres)
            palabras = re.findall(r'\b[a-zA-ZáéíóúüñÁÉÍÓÚÜÑ][a-zA-ZáéíóúüñÁÉÍÓÚÜÑ0-9]*\b', contenido)
            # Filtrar palabras muy cortas (opcional)
            palabras = [p.lower() for p in palabras if len(p) >= 2]
            return palabras
        
        elif self.tipo_dato == "lineas":
            # Separar por líneas
            lineas = contenido.split('\n')
            # Filtrar líneas vacías
            lineas = [l.strip() for l in lineas if l.strip()]
            return lineas
        
        elif self.tipo_dato == "fechas":
            # Buscar patrones de fechas
            patrones_fecha = [
                r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # DD/MM/YYYY o MM/DD/YYYY
                r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',    # YYYY-MM-DD
                r'\d{1,2}\s+(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+\d{4}',
                r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}'
            ]
            
            fechas = []
            for patron in patrones_fecha:
                encontradas = re.findall(patron, contenido, re.IGNORECASE)
                fechas.extend(encontradas)
            
            # Intentar convertir a objetos datetime para ordenamiento
            fechas_convertidas = []
            for f in fechas:
                try:
                    # Intentar diferentes formatos
                    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y']:
                        try:
                            fecha_obj = datetime.strptime(f, fmt)
                            fechas_convertidas.append(fecha_obj)
                            break
                        except:
                            continue
                    else:
                        fechas_convertidas.append(f)
                except:
                    fechas_convertidas.append(f)
            
            return fechas_convertidas
        
        elif self.tipo_dato == "caracteres":
            # Extraer caracteres individuales (excluyendo espacios y saltos de línea)
            caracteres = [c for c in contenido if c.strip() and not c.isspace()]
            return caracteres
        
        else:  # "todo"
            # Dividir en elementos (por espacios, comas, etc.)
            elementos = re.findall(r'[^\s,;:!?¡¿()\[\]{}<>]+', contenido)
            return elementos
    
    def mostrar_datos_en_canvas(self):
        """Mostrar los datos de ambos archivos en el canvas (con scroll)"""
        # Limpiar frame interno
        for widget in self.datos_inner_frame.winfo_children():
            widget.destroy()
        
        if not self.datos_archivo1 and not self.datos_archivo2:
            return
        
        y_pos = 10
        
        if self.datos_archivo1:
            lbl1 = tk.Label(self.datos_inner_frame, 
                           text=f"Archivo 1 ({self.nombre_archivo1}): {len(self.datos_archivo1)} elementos", 
                           fg="#3498db", font=("Arial", 10, "bold"), bg="white")
            lbl1.pack(anchor="w", pady=(0,5))
            
            # Mostrar elementos en un Text con scroll
            frame1 = tk.Frame(self.datos_inner_frame, bg="white")
            frame1.pack(fill=tk.X, pady=(0,10))
            
            text1 = tk.Text(frame1, font=("Courier", 9), height=min(10, len(self.datos_archivo1)//10 + 2), 
                           wrap=tk.WORD, bg="#f8f9fa")
            scroll1 = tk.Scrollbar(frame1, orient="vertical", command=text1.yview)
            text1.configure(yscrollcommand=scroll1.set)
            
            text1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scroll1.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Insertar elementos
            texto = ""
            for i, elem in enumerate(self.datos_archivo1[:200]):  # Limitar a 200 para rendimiento
                if self.tipo_dato == "numeros":
                    texto += f"{elem}, "
                else:
                    # Truncar strings largos
                    elem_str = str(elem)[:50]
                    texto += f"{elem_str}, "
                if (i + 1) % 10 == 0:
                    texto += "\n"
            
            if len(self.datos_archivo1) > 200:
                texto += f"\n... y {len(self.datos_archivo1)-200} elementos más"
            
            text1.insert(tk.END, texto.strip(", \n"))
            text1.config(state=tk.DISABLED)
        
        if self.datos_archivo2:
            lbl2 = tk.Label(self.datos_inner_frame, 
                           text=f"Archivo 2 ({self.nombre_archivo2}): {len(self.datos_archivo2)} elementos", 
                           fg="#e74c3c", font=("Arial", 10, "bold"), bg="white")
            lbl2.pack(anchor="w", pady=(10,5))
            
            frame2 = tk.Frame(self.datos_inner_frame, bg="white")
            frame2.pack(fill=tk.X)
            
            text2 = tk.Text(frame2, font=("Courier", 9), height=min(10, len(self.datos_archivo2)//10 + 2), 
                           wrap=tk.WORD, bg="#f8f9fa")
            scroll2 = tk.Scrollbar(frame2, orient="vertical", command=text2.yview)
            text2.configure(yscrollcommand=scroll2.set)
            
            text2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scroll2.pack(side=tk.RIGHT, fill=tk.Y)
            
            texto = ""
            for i, elem in enumerate(self.datos_archivo2[:200]):
                if self.tipo_dato == "numeros":
                    texto += f"{elem}, "
                else:
                    elem_str = str(elem)[:50]
                    texto += f"{elem_str}, "
                if (i + 1) % 10 == 0:
                    texto += "\n"
            
            if len(self.datos_archivo2) > 200:
                texto += f"\n... y {len(self.datos_archivo2)-200} elementos más"
            
            text2.insert(tk.END, texto.strip(", \n"))
            text2.config(state=tk.DISABLED)
    
    def comparar_archivos(self):
        """Comparar los dos archivos cargados"""
        if not self.datos_archivo1 or not self.datos_archivo2:
            messagebox.showwarning("Advertencia", "Debe cargar ambos archivos primero.")
            return
        
        # Para comparación, convertir todo a string si no son números
        if self.tipo_dato == "numeros":
            set1 = set(self.datos_archivo1)
            set2 = set(self.datos_archivo2)
        else:
            set1 = set(str(x) for x in self.datos_archivo1)
            set2 = set(str(x) for x in self.datos_archivo2)
        
        comunes = set1 & set2
        solo1 = set1 - set2
        solo2 = set2 - set1
        
        # Estadísticas según tipo
        if self.tipo_dato == "numeros":
            stats1 = {
                "min": min(self.datos_archivo1),
                "max": max(self.datos_archivo1),
                "suma": sum(self.datos_archivo1),
                "promedio": sum(self.datos_archivo1) / len(self.datos_archivo1)
            }
            stats2 = {
                "min": min(self.datos_archivo2),
                "max": max(self.datos_archivo2),
                "suma": sum(self.datos_archivo2),
                "promedio": sum(self.datos_archivo2) / len(self.datos_archivo2)
            }
        else:
            stats1 = {
                "min": "N/A",
                "max": "N/A", 
                "suma": "N/A",
                "promedio": "N/A"
            }
            stats2 = {
                "min": "N/A",
                "max": "N/A",
                "suma": "N/A", 
                "promedio": "N/A"
            }
        
        self.resultados_comparacion = {
            "archivo1": {
                "nombre": self.nombre_archivo1,
                "elementos": len(self.datos_archivo1),
                "unicos": len(set1),
                **stats1
            },
            "archivo2": {
                "nombre": self.nombre_archivo2,
                "elementos": len(self.datos_archivo2),
                "unicos": len(set2),
                **stats2
            },
            "comparacion": {
                "elementos_comunes": len(comunes),
                "comunes": sorted(list(comunes))[:50] if comunes else [],
                "solo_archivo1": len(solo1),
                "solo_archivo1_valores": sorted(list(solo1))[:50] if solo1 else [],
                "solo_archivo2": len(solo2),
                "solo_archivo2_valores": sorted(list(solo2))[:50] if solo2 else [],
                "similitud": (len(comunes) / max(len(set1), len(set2)) * 100) if max(len(set1), len(set2)) > 0 else 0
            },
            "tipo_dato": self.tipo_dato
        }
        
        self.mostrar_comparacion()
    
    def mostrar_comparacion(self):
        """Mostrar ventana de comparación"""
        comp_win = tk.Toplevel(self.root)
        comp_win.title("Comparación de Archivos")
        comp_win.geometry("800x700")
        comp_win.configure(bg="#ecf0f1")
        
        tk.Label(comp_win, text="COMPARACIÓN DE ARCHIVOS", 
                font=("Arial", 14, "bold"), bg="#2c3e50", fg="white", pady=10).pack(fill=tk.X)
        
        text_frame = tk.Frame(comp_win, bg="#ecf0f1")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        texto = tk.Text(text_frame, font=("Courier", 10), wrap=tk.WORD, 
                       yscrollcommand=scrollbar.set, padx=15, pady=15)
        texto.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=texto.yview)
        
        # Escribir resultados
        texto.insert(tk.END, "="*70 + "\n")
        texto.insert(tk.END, f"TIPO DE DATO ANALIZADO: {self.resultados_comparacion['tipo_dato'].upper()}\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, "ESTADÍSTICAS DE CADA ARCHIVO\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, f"ARCHIVO 1: {self.resultados_comparacion['archivo1']['nombre']}\n")
        texto.insert(tk.END, f"  • Elementos: {self.resultados_comparacion['archivo1']['elementos']:,}\n")
        texto.insert(tk.END, f"  • Valores únicos: {self.resultados_comparacion['archivo1']['unicos']:,}\n")
        
        if self.tipo_dato == "numeros":
            texto.insert(tk.END, f"  • Mínimo: {self.resultados_comparacion['archivo1']['min']}\n")
            texto.insert(tk.END, f"  • Máximo: {self.resultados_comparacion['archivo1']['max']}\n")
            texto.insert(tk.END, f"  • Suma: {self.resultados_comparacion['archivo1']['suma']:,}\n")
            texto.insert(tk.END, f"  • Promedio: {self.resultados_comparacion['archivo1']['promedio']:.2f}\n")
        
        texto.insert(tk.END, "\n")
        texto.insert(tk.END, f"ARCHIVO 2: {self.resultados_comparacion['archivo2']['nombre']}\n")
        texto.insert(tk.END, f"  • Elementos: {self.resultados_comparacion['archivo2']['elementos']:,}\n")
        texto.insert(tk.END, f"  • Valores únicos: {self.resultados_comparacion['archivo2']['unicos']:,}\n")
        
        if self.tipo_dato == "numeros":
            texto.insert(tk.END, f"  • Mínimo: {self.resultados_comparacion['archivo2']['min']}\n")
            texto.insert(tk.END, f"  • Máximo: {self.resultados_comparacion['archivo2']['max']}\n")
            texto.insert(tk.END, f"  • Suma: {self.resultados_comparacion['archivo2']['suma']:,}\n")
            texto.insert(tk.END, f"  • Promedio: {self.resultados_comparacion['archivo2']['promedio']:.2f}\n")
        
        texto.insert(tk.END, "\n" + "="*70 + "\n")
        texto.insert(tk.END, "COMPARACIÓN ENTRE ARCHIVOS\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, f"📊 Elementos comunes: {self.resultados_comparacion['comparacion']['elementos_comunes']:,}\n")
        texto.insert(tk.END, f"📊 Similitud: {self.resultados_comparacion['comparacion']['similitud']:.2f}%\n")
        texto.insert(tk.END, f"🔴 Solo en Archivo 1: {self.resultados_comparacion['comparacion']['solo_archivo1']:,}\n")
        texto.insert(tk.END, f"🔵 Solo en Archivo 2: {self.resultados_comparacion['comparacion']['solo_archivo2']:,}\n\n")
        
        if self.resultados_comparacion['comparacion']['elementos_comunes'] > 0:
            texto.insert(tk.END, f"Elementos comunes (primeros 50):\n")
            texto.insert(tk.END, f"{self.resultados_comparacion['comparacion']['comunes']}\n\n")
        
        if self.resultados_comparacion['comparacion']['solo_archivo1'] > 0:
            texto.insert(tk.END, f"Elementos solo en Archivo 1 (primeros 50):\n")
            texto.insert(tk.END, f"{self.resultados_comparacion['comparacion']['solo_archivo1_valores']}\n\n")
        
        if self.resultados_comparacion['comparacion']['solo_archivo2'] > 0:
            texto.insert(tk.END, f"Elementos solo en Archivo 2 (primeros 50):\n")
            texto.insert(tk.END, f"{self.resultados_comparacion['comparacion']['solo_archivo2_valores']}\n\n")
        
        texto.config(state=tk.DISABLED)
        
        btn_cerrar = tk.Button(comp_win, text="CERRAR", command=comp_win.destroy,
                              bg="#2c3e50", fg="white", font=("Arial", 10, "bold"),
                              width=15)
        btn_cerrar.pack(pady=10)
    
    def combinar_y_ordenar(self):
        """Combinar los dos archivos y ordenarlos"""
        if not self.datos_archivo1 or not self.datos_archivo2:
            messagebox.showwarning("Advertencia", "Debe cargar ambos archivos primero.")
            return
        
        # Combinar datos
        self.datos_combinados = self.datos_archivo1 + self.datos_archivo2
        
        # Ordenar según el tipo de dato
        if self.tipo_dato == "numeros":
            self.datos_ejemplo = sorted(self.datos_combinados)
        else:
            # Para strings, ordenar alfabéticamente
            self.datos_ejemplo = sorted(self.datos_combinados, key=str)
        
        # Mostrar información de combinación
        self.lbl_combinacion.config(text=f"📊 COMBINACIÓN: {len(self.datos_archivo1)} + {len(self.datos_archivo2)} = {len(self.datos_combinados)} elementos totales - Tipo: {self.tipo_dato}")
        
        # Iniciar ordenamiento
        self.iniciar_ordenamiento_simultaneo()
    
    def crear_panel_algoritmo(self, columna, titulo, color):
        frame = tk.LabelFrame(self.algoritmos_inner_frame, text=titulo, font=("Arial", 11, "bold"),
                             bg="#ecf0f1", fg=color, bd=2, relief=tk.RAISED)
        frame.grid(row=0, column=columna, padx=10, pady=10, sticky="nsew")
        
        # Canvas para dibujar con scroll interno
        canvas_container = tk.Frame(frame, bg="white")
        canvas_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Agregar scroll al canvas de dibujo
        canvas_scrollbar = tk.Scrollbar(canvas_container, orient="vertical")
        canvas = tk.Canvas(canvas_container, bg="white", height=300, 
                          yscrollcommand=canvas_scrollbar.set, highlightthickness=1)
        canvas_scrollbar.config(command=canvas.yview)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        label_datos = tk.Label(frame, text="", bg="#ecf0f1", font=("Courier", 8), wraplength=380)
        label_datos.pack(fill=tk.X, padx=5, pady=5)
        
        label_stats = tk.Label(frame, text="Comparaciones: 0 | Tiempo: 0 ms", 
                              bg="#ecf0f1", font=("Arial", 9), fg="#2c3e50")
        label_stats.pack(fill=tk.X, padx=5, pady=2)
        
        label_estado = tk.Label(frame, text="⚡ En espera", bg="#ecf0f1", font=("Arial", 9, "italic"), fg="#7f8c8d")
        label_estado.pack(fill=tk.X, padx=5, pady=2)
        
        frame.canvas = canvas
        frame.canvas_container = canvas_container
        frame.canvas_scrollbar = canvas_scrollbar
        frame.label_datos = label_datos
        frame.label_stats = label_stats
        frame.label_estado = label_estado
        frame.color = color
        
        return frame
    
    def dibujar_barras(self, canvas, datos, color_base):
        canvas.delete("all")
        if not datos or self.tipo_dato != "numeros":
            # Si no son números, mostrar mensaje
            canvas.create_text(200, 150, text="Visualización disponible solo para datos numéricos", 
                              fill="gray", font=("Arial", 10))
            return
        
        width = canvas.winfo_width() if canvas.winfo_width() > 50 else 380
        height = canvas.winfo_height() if canvas.winfo_height() > 50 else 300
        
        n = min(len(datos), 100)  # Limitar a 100 elementos para visualización
        if n == 0:
            return
        
        bar_width = max(2, (width - 20) / n - 2)
        max_valor = max(datos[:n]) if datos else 1
        min_valor = min(datos[:n]) if datos else 0
        rango = max_valor - min_valor if max_valor != min_valor else 1
        
        total_width = n * (bar_width + 2)
        offset = max(5, (width - total_width) / 2)
        
        for i, valor in enumerate(datos[:n]):
            x0 = offset + i * (bar_width + 2)
            bar_height = ((valor - min_valor) / rango) * (height - 60)
            y0 = height - bar_height - 40
            y1 = height - 40
            
            intensidad = (valor - min_valor) / rango
            r = int(255 * (1 - intensidad))
            g = int(255 * intensidad)
            b = 100
            color = f"#{r:02x}{g:02x}{b:02x}"
            
            canvas.create_rectangle(x0, y0, x0 + bar_width, y1, fill=color, outline="#34495e", width=1)
            
            if bar_width >= 12 and n <= 30:
                canvas.create_text(x0 + bar_width/2, y1 - 5, text=str(valor), anchor="n", font=("Arial", 7))
            elif bar_width >= 8 and n <= 60:
                canvas.create_text(x0 + bar_width/2, y1 - 5, text=str(valor), anchor="n", font=("Arial", 6))
        
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.update()
    
    def actualizar_visualizacion(self, frame, datos, mensaje, comparaciones, tiempo_ms):
        if self.tipo_dato == "numeros":
            datos_str = str(datos[:30]) if datos else "[]"
            if len(datos) > 30:
                datos_str = datos_str[:-1] + f", ... +{len(datos)-30} más]"
        else:
            datos_str = f"[{len(datos)} elementos]"
        
        frame.label_datos.config(text=f"Datos: {datos_str}")
        frame.label_stats.config(text=f"Comparaciones: {comparaciones} | Tiempo: {tiempo_ms:.0f} ms")
        frame.label_estado.config(text=mensaje)
        self.dibujar_barras(frame.canvas, datos, frame.color)
    
    def mezcla_directa(self, frame, datos_originales, resultados):
        inicio = time.time()
        datos = datos_originales.copy()
        paso = 1
        fase = 1
        comparaciones = 0
        
        self.actualizar_visualizacion(frame, datos, f"Fase {fase}: Inicio", comparaciones, 0)
        time.sleep(0.05)
        
        while paso < len(datos):
            fase += 1
            for i in range(0, len(datos), paso * 2):
                izquierda = datos[i:i + paso]
                derecha = datos[i + paso:i + paso * 2]
                mezclados = []
                a, b = 0, 0
                while a < len(izquierda) and b < len(derecha):
                    comparaciones += 1
                    # Usar el comparador según tipo de dato
                    if self.comparar_elementos(izquierda[a], derecha[b]) < 0:
                        mezclados.append(izquierda[a])
                        a += 1
                    else:
                        mezclados.append(derecha[b])
                        b += 1
                mezclados.extend(izquierda[a:])
                mezclados.extend(derecha[b:])
                datos[i:i + paso * 2] = mezclados
            
            self.actualizar_visualizacion(frame, datos, f"Fase {fase}: bloque tamaño {paso * 2}", comparaciones, (time.time() - inicio) * 1000)
            time.sleep(0.05)
            paso *= 2
        
        tiempo_ms = (time.time() - inicio) * 1000
        self.actualizar_visualizacion(frame, datos, "✅ COMPLETADO", comparaciones, tiempo_ms)
        resultados["mezcla_directa"] = {"datos": datos, "comparaciones": comparaciones, "tiempo": tiempo_ms}
        return datos
    
    def comparar_elementos(self, a, b):
        """Comparar dos elementos según el tipo de dato seleccionado"""
        if self.tipo_dato == "numeros":
            if a < b:
                return -1
            elif a > b:
                return 1
            return 0
        elif self.tipo_dato == "fechas":
            # Si son objetos datetime
            if hasattr(a, '__lt__') and hasattr(b, '__lt__'):
                if a < b:
                    return -1
                elif a > b:
                    return 1
                return 0
        else:
            # Strings
            str_a = str(a).lower()
            str_b = str(b).lower()
            if str_a < str_b:
                return -1
            elif str_a > str_b:
                return 1
            return 0
    
    def mezcla_equilibrada_recursiva(self, frame, datos, profundidad=0, comparaciones=[0]):
        if len(datos) <= 1:
            return datos
        
        medio = len(datos) // 2
        izquierda = self.mezcla_equilibrada_recursiva(frame, datos[:medio], profundidad + 1, comparaciones)
        derecha = self.mezcla_equilibrada_recursiva(frame, datos[medio:], profundidad + 1, comparaciones)
        
        resultado = []
        i = j = 0
        while i < len(izquierda) and j < len(derecha):
            comparaciones[0] += 1
            if self.comparar_elementos(izquierda[i], derecha[j]) < 0:
                resultado.append(izquierda[i])
                i += 1
            else:
                resultado.append(derecha[j])
                j += 1
        resultado.extend(izquierda[i:])
        resultado.extend(derecha[j:])
        
        return resultado
    
    def mezcla_equilibrada(self, frame, datos_originales, resultados):
        inicio = time.time()
        frame.label_estado.config(text="Dividiendo lista...")
        time.sleep(0.05)
        
        comparaciones = [0]
        frame.label_estado.config(text="Mezclando niveles...")
        time.sleep(0.05)
        
        datos_ordenados = self.mezcla_equilibrada_recursiva(frame, datos_originales, 0, comparaciones)
        self.actualizar_visualizacion(frame, datos_ordenados, "Mezclando...", comparaciones[0], (time.time() - inicio) * 1000)
        time.sleep(0.05)
        
        tiempo_ms = (time.time() - inicio) * 1000
        self.actualizar_visualizacion(frame, datos_ordenados, "✅ COMPLETADO", comparaciones[0], tiempo_ms)
        resultados["mezcla_equilibrada"] = {"datos": datos_ordenados, "comparaciones": comparaciones[0], "tiempo": tiempo_ms}
        return datos_ordenados
    
    def intercalacion(self, frame, datos_originales, resultados):
        inicio = time.time()
        datos = datos_originales.copy()
        k = 2
        fase = 1
        comparaciones = 0
        
        self.actualizar_visualizacion(frame, datos, f"Fase {fase}: bloque tamaño 1", comparaciones, 0)
        time.sleep(0.05)
        
        while k <= len(datos):
            fase += 1
            for i in range(0, len(datos), k):
                sublista = datos[i:i + k]
                # Ordenar usando el comparador
                sublista.sort(key=lambda x: str(x) if self.tipo_dato != "numeros" else x)
                comparaciones += len(sublista) * (len(sublista) - 1) // 2
                datos[i:i + k] = sublista
            
            self.actualizar_visualizacion(frame, datos, f"Fase {fase}: bloque tamaño {k}", comparaciones, (time.time() - inicio) * 1000)
            time.sleep(0.05)
            
            if k >= len(datos):
                break
            k *= 2
        
        if k > len(datos) and len(datos) > 2:
            datos.sort(key=lambda x: str(x) if self.tipo_dato != "numeros" else x)
            self.actualizar_visualizacion(frame, datos, "Fase final: ordenando resto", comparaciones, (time.time() - inicio) * 1000)
            time.sleep(0.05)
        
        tiempo_ms = (time.time() - inicio) * 1000
        self.actualizar_visualizacion(frame, datos, "✅ COMPLETADO", comparaciones, tiempo_ms)
        resultados["intercalacion"] = {"datos": datos, "comparaciones": comparaciones, "tiempo": tiempo_ms}
        return datos
    
    def iniciar_ordenamiento_simultaneo(self):
        if self.ejecutando or not self.datos_ejemplo:
            return
        
        self.ejecutando = True
        self.progress.start()
        
        self.resultados = {}
        
        for frame in [self.frame_mezcla_directa, self.frame_mezcla_equilibrada, self.frame_intercalacion]:
            frame.label_estado.config(text="🔄 En ejecución...")
            frame.label_stats.config(text="Comparaciones: 0 | Tiempo: 0 ms")
        
        hilo1 = threading.Thread(target=self.mezcla_directa, args=(self.frame_mezcla_directa, self.datos_ejemplo, self.resultados))
        hilo2 = threading.Thread(target=self.mezcla_equilibrada, args=(self.frame_mezcla_equilibrada, self.datos_ejemplo, self.resultados))
        hilo3 = threading.Thread(target=self.intercalacion, args=(self.frame_intercalacion, self.datos_ejemplo, self.resultados))
        
        hilo1.start()
        hilo2.start()
        hilo3.start()
        
        self.monitor_hilos(hilo1, hilo2, hilo3)
    
    def monitor_hilos(self, hilo1, hilo2, hilo3):
        if not (hilo1.is_alive() or hilo2.is_alive() or hilo3.is_alive()):
            self.ejecutando = False
            self.progress.stop()
            self.mostrar_resumen_combinado()
        else:
            self.root.after(500, lambda: self.monitor_hilos(hilo1, hilo2, hilo3))
    
    def mostrar_resumen_combinado(self):
        """Mostrar resumen de los datos combinados y ordenados"""
        resumen_win = tk.Toplevel(self.root)
        resumen_win.title("Resultados - Datos Combinados y Ordenados")
        resumen_win.geometry("900x700")
        resumen_win.configure(bg="#ecf0f1")
        
        tk.Label(resumen_win, text="RESULTADOS - DATOS COMBINADOS Y ORDENADOS", 
                font=("Arial", 14, "bold"), bg="#2c3e50", fg="white", pady=10).pack(fill=tk.X)
        
        text_frame = tk.Frame(resumen_win, bg="#ecf0f1")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        texto = tk.Text(text_frame, font=("Courier", 10), wrap=tk.WORD, 
                       yscrollcommand=scrollbar.set, padx=15, pady=15)
        texto.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=texto.yview)
        
        texto.insert(tk.END, "="*70 + "\n")
        texto.insert(tk.END, f"TIPO DE DATO ANALIZADO: {self.tipo_dato.upper()}\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, "INFORMACIÓN DE LOS ARCHIVOS ORIGINALES\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, f"📁 ARCHIVO 1: {self.nombre_archivo1}\n")
        texto.insert(tk.END, f"   Elementos: {len(self.datos_archivo1):,}\n")
        if self.tipo_dato == "numeros":
            texto.insert(tk.END, f"   Rango: [{min(self.datos_archivo1)}, {max(self.datos_archivo1)}]\n")
        texto.insert(tk.END, "\n")
        
        texto.insert(tk.END, f"📁 ARCHIVO 2: {self.nombre_archivo2}\n")
        texto.insert(tk.END, f"   Elementos: {len(self.datos_archivo2):,}\n")
        if self.tipo_dato == "numeros":
            texto.insert(tk.END, f"   Rango: [{min(self.datos_archivo2)}, {max(self.datos_archivo2)}]\n")
        texto.insert(tk.END, "\n")
        
        texto.insert(tk.END, "="*70 + "\n")
        texto.insert(tk.END, "DATOS COMBINADOS\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        texto.insert(tk.END, f"📊 Total elementos combinados: {len(self.datos_combinados):,}\n")
        texto.insert(tk.END, f"📊 Elementos únicos: {len(set(str(x) for x in self.datos_combinados)):,}\n")
        
        if self.tipo_dato == "numeros":
            texto.insert(tk.END, f"📊 Rango completo: [{min(self.datos_combinados)}, {max(self.datos_combinados)}]\n")
            texto.insert(tk.END, f"📊 Suma total: {sum(self.datos_combinados):,}\n")
            texto.insert(tk.END, f"📊 Promedio: {sum(self.datos_combinados)/len(self.datos_combinados):.2f}\n")
        
        texto.insert(tk.END, "\n" + "="*70 + "\n")
        texto.insert(tk.END, "RESULTADOS DE ORDENAMIENTO\n")
        texto.insert(tk.END, "="*70 + "\n\n")
        
        algoritmos = [
            ("MEZCLA DIRECTA", self.resultados.get("mezcla_directa", {})),
            ("MEZCLA EQUILIBRADA", self.resultados.get("mezcla_equilibrada", {})),
            ("INTERCALACIÓN", self.resultados.get("intercalacion", {}))
        ]
        
        for nombre, res in algoritmos:
            if res:
                texto.insert(tk.END, f"{nombre}\n")
                texto.insert(tk.END, "-"*40 + "\n")
                texto.insert(tk.END, f"  • Comparaciones: {res['comparaciones']:,}\n")
                texto.insert(tk.END, f"  • Tiempo: {res['tiempo']:.2f} ms\n")
                texto.insert(tk.END, f"  • Verificación: ✓ COMPLETADO\n\n")
        
        # Mostrar primeros elementos del resultado
        if self.resultados:
            primer_algoritmo = list(self.resultados.values())[0]
            datos_ordenados = primer_algoritmo['datos']
            
            texto.insert(tk.END, "="*70 + "\n")
            texto.insert(tk.END, "MUESTRA DEL RESULTADO ORDENADO\n")
            texto.insert(tk.END, "="*70 + "\n\n")
            
            if len(datos_ordenados) > 50:
                texto.insert(tk.END, f"Primeros 50 elementos:\n")
                for i, elem in enumerate(datos_ordenados[:50], 1):
                    texto.insert(tk.END, f"{i}. {str(elem)[:100]}\n")
                texto.insert(tk.END, f"\n... y {len(datos_ordenados)-50} elementos más\n\n")
            else:
                texto.insert(tk.END, f"Todos los elementos:\n")
                for i, elem in enumerate(datos_ordenados, 1):
                    texto.insert(tk.END, f"{i}. {str(elem)[:100]}\n")
        
        texto.config(state=tk.DISABLED)
        
        btn_frame = tk.Frame(resumen_win, bg="#ecf0f1", pady=10)
        btn_frame.pack()
        
        tk.Button(btn_frame, text="GUARDAR RESULTADO", command=self.guardar_resultado_combinado,
                 bg="#27ae60", fg="white", font=("Arial", 10, "bold"),
                 width=15).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="CERRAR", command=resumen_win.destroy,
                 bg="#2c3e50", fg="white", font=("Arial", 10, "bold"),
                 width=15).pack(side=tk.LEFT, padx=5)
    
    def guardar_resultado_combinado(self):
        """Guardar el resultado combinado y ordenado en un archivo"""
        if not self.resultados:
            messagebox.showwarning("Advertencia", "Primero debe combinar y ordenar los datos.")
            return
        
        formatos = [
            ("Archivo de texto", "*.txt"),
            ("Archivo CSV", "*.csv"),
            ("Archivo JSON", "*.json"),
            ("Archivo HTML", "*.html"),
        ]
        
        ruta = filedialog.asksaveasfilename(
            title="Guardar resultado combinado",
            defaultextension=".txt",
            filetypes=formatos
        )
        
        if not ruta:
            return
        
        try:
            extension = Path(ruta).suffix.lower()
            datos_ordenados = list(self.resultados.values())[0]['datos'] if self.resultados else []
            
            if extension == '.txt':
                with open(ruta, 'w', encoding='utf-8') as f:
                    f.write("="*70 + "\n")
                    f.write("RESULTADO DE COMBINACIÓN Y ORDENAMIENTO\n")
                    f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Tipo de dato: {self.tipo_dato}\n")
                    f.write("="*70 + "\n\n")
                    
                    f.write(f"Archivo 1: {self.nombre_archivo1} ({len(self.datos_archivo1)} elementos)\n")
                    f.write(f"Archivo 2: {self.nombre_archivo2} ({len(self.datos_archivo2)} elementos)\n")
                    f.write(f"Total combinado: {len(self.datos_combinados)} elementos\n\n")
                    
                    f.write("DATOS ORDENADOS:\n")
                    f.write("-"*40 + "\n")
                    for i, elem in enumerate(datos_ordenados, 1):
                        f.write(f"{i}. {elem}\n")
            
            elif extension == '.csv':
                with open(ruta, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Indice", "Valor"])
                    for i, elem in enumerate(datos_ordenados, 1):
                        writer.writerow([i, elem])
            
            elif extension == '.json':
                with open(ruta, 'w', encoding='utf-8') as f:
                    # Convertir elementos a string si es necesario
                    datos_serializables = [str(elem) if not isinstance(elem, (int, float)) else elem for elem in datos_ordenados]
                    json.dump({
                        "metadata": {
                            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "tipo_dato": self.tipo_dato,
                            "archivo1": self.nombre_archivo1,
                            "archivo2": self.nombre_archivo2,
                            "elementos_archivo1": len(self.datos_archivo1),
                            "elementos_archivo2": len(self.datos_archivo2),
                            "total_combinado": len(self.datos_combinados)
                        },
                        "datos_ordenados": datos_serializables
                    }, f, indent=2, ensure_ascii=False)
            
            elif extension == '.html':
                with open(ruta, 'w', encoding='utf-8') as f:
                    f.write(f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Resultados Combinados</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
.container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; }}
h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
h2 {{ color: #34495e; margin-top: 20px; }}
.stats {{ background: #ecf0f1; padding: 15px; border-radius: 5px; margin: 10px 0; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
th {{ background-color: #2c3e50; color: white; }}
tr:nth-child(even) {{ background-color: #f2f2f2; }}
.scrollable {{ max-height: 400px; overflow-y: auto; }}
</style>
</head>
<body>
<div class="container">
<h1>📊 Resultados de Combinación y Ordenamiento</h1>
<p><strong>Fecha:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
<p><strong>Tipo de dato:</strong> {self.tipo_dato}</p>

<h2>📁 Archivos Fuente</h2>
<div class="stats">
<p><strong>Archivo 1:</strong> {self.nombre_archivo1} ({len(self.datos_archivo1)} elementos)</p>
<p><strong>Archivo 2:</strong> {self.nombre_archivo2} ({len(self.datos_archivo2)} elementos)</p>
<p><strong>Total combinado:</strong> {len(self.datos_combinados)} elementos</p>
</div>

<h2>📈 Estadísticas</h2>
<div class="stats">
<p><strong>Elementos únicos:</strong> {len(set(str(x) for x in self.datos_combinados))}</p>
</div>

<h2>🔢 Datos Ordenados</h2>
<div class="scrollable">
<table>
<tr><th>Posición</th><th>Valor</th></tr>
""")
                    
                    for i, elem in enumerate(datos_ordenados[:500], 1):
                        f.write(f"<tr><td>{i}</td><td>{elem}</td></tr>\n")
                    
                    if len(datos_ordenados) > 500:
                        f.write(f"<tr><td colspan='2'>... y {len(datos_ordenados)-500} elementos más ...</td></tr>\n")
                    
                    f.write("""</table>
</div>
</div>
</body>
</html>""")
            
            messagebox.showinfo("Éxito", f"Archivo combinado guardado correctamente:\n{ruta}\n\n{len(datos_ordenados)} elementos ordenados")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = OrdenamientoExternoSimultaneo(root)
    
    messagebox.showinfo("Bienvenido - Comparación de Dos Archivos", 
                       "📊 PROGRAMA DE COMPARACIÓN Y ORDENAMIENTO\n\n"
                       "NUEVAS FUNCIONALIDADES:\n"
                       "✓ SELECCIÓN DE HOJAS DE EXCEL:\n"
                       "  • Al cargar un archivo Excel, podrá elegir qué hojas procesar\n"
                       "  • Selección múltiple con opción de 'Seleccionar todas' y 'Ninguna'\n"
                       "  • Las hojas seleccionadas se combinan y procesan juntas\n"
                       "✓ SELECCIÓN DE TIPO DE DATO:\n"
                       "  • 🔢 Números: Extrae y ordena valores numéricos\n"
                       "  • 📝 Palabras: Extrae y ordena palabras\n"
                       "  • 📄 Líneas: Ordena líneas completas\n"
                       "  • 📅 Fechas: Detecta y ordena fechas\n"
                       "  • 🔤 Caracteres: Ordena caracteres individuales\n"
                       "  • 🌐 Todo: Procesa todo el texto\n"
                       "✓ Scroll vertical en toda la interfaz\n"
                       "✓ Vista previa con scroll para datos grandes\n"
                       "✓ Navegación con rueda del mouse\n\n"
                       "Instrucciones:\n"
                       "1. Seleccione el tipo de dato a analizar\n"
                       "2. Cargue Archivo 1 y Archivo 2\n"
                       "3. Para archivos Excel, elija las hojas que desea procesar\n"
                       "4. Use 'COMPARAR ARCHIVOS' para ver diferencias\n"
                       "5. Use 'COMBINAR Y ORDENAR' para unificar y ordenar\n"
                       "6. Guarde el resultado ordenado\n\n"
                       "💡 Use la rueda del mouse o la barra lateral para desplazarse")
    
    root.mainloop()